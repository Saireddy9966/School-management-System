from django.http import JsonResponse
import json
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, login, logout
from django.core.mail import send_mail
from django.db import transaction
from django.http import HttpResponseServerError
from django.contrib import messages
from django.http import HttpResponseBadRequest
from django.shortcuts import HttpResponse
from django.db.models import F
from functools import wraps
from django.contrib.auth.models import User, Group
from django.core.serializers import serialize
from .forms import *
from .models import *
from .utils import *
import logging
from django.utils import timezone
from datetime import date
from collections import defaultdict
from django.db.models import Sum, Count,Q
from django.db.models.functions import TruncMonth,ExtractMonth, ExtractYear
import calendar
from datetime import datetime
import openpyxl
import csv
from openpyxl.utils.exceptions import InvalidFileException
import os
import pandas as pd
from io import TextIOWrapper
from io import BytesIO
from django.conf import settings
from twilio.rest import Client
import razorpay
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

logger = logging.getLogger(__name__)

# Custom decorators
def superuser_required(view_func):
    def check_user_is_superuser(user):
        return user.is_authenticated and user.is_superuser

    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return login_required(lambda x: x)(request, *args, **kwargs)
        
        if not check_user_is_superuser(request.user):
            messages.error(request, "You do not have Admin permissions to access this page.")
            return redirect('home') 

        return view_func(request, *args, **kwargs)
    return login_required(wrapped_view)


def staff_required(view_func):
    def check_user_is_staff(user):
        return user.is_authenticated and user.groups.filter(name='Staff').exists()

    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return login_required(lambda x: x)(request, *args, **kwargs)
        
        if not check_user_is_staff(request.user):
            messages.error(request, "You do not have Staff permission to access this page.")
            return redirect('home')

        return view_func(request, *args, **kwargs)
    return login_required(wrapped_view)

def faculty_required(view_func):
    def check_user_is_staff(user):
        return user.is_authenticated and user.groups.filter(name='Faculty').exists()

    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return login_required(lambda x: x)(request, *args, **kwargs)
        
        if not check_user_is_staff(request.user):
            messages.error(request, "You do not have Faculty permission to access this page.")
            return redirect('home')

        return view_func(request, *args, **kwargs)
    return login_required(wrapped_view)

def student_required(view_func):
    def check_user_is_staff(user):
        return user.is_authenticated and user.groups.filter(name='Student').exists()

    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return login_required(lambda x: x)(request, *args, **kwargs)
        
        if not check_user_is_staff(request.user):
            messages.error(request, "You do not have Student permission to access this page.")
            return redirect('home')

        return view_func(request, *args, **kwargs)
    return login_required(wrapped_view)

def parent_required(view_func):
    def check_user_is_staff(user):
        return user.is_authenticated and user.groups.filter(name='Parent').exists()

    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return login_required(lambda x: x)(request, *args, **kwargs)
        
        if not check_user_is_staff(request.user):
            messages.error(request, "You do not have Parent permission to access this page.")
            return redirect('home')

        return view_func(request, *args, **kwargs)
    return login_required(wrapped_view)

# view function 
# Home

def home_view(request):
    sections = Section.objects.all()
    registration_form = RegistrationForm()
    staff_form = StaffForm()
    faculty_form = FacultyForm()
    students_form = StudentsForm()
    parents_form = ParentsForm()
    
    if request.method == "POST":
        if 'login' in request.POST:
            username = request.POST.get("username")
            password = request.POST.get("password")

            if not User.objects.filter(username=username).exists():
                messages.error(request, "This username does not exist")
                return redirect("home")

            user = authenticate(username=username, password=password)

            if user is None:
                messages.error(request, "Invalid Credential")
                return redirect("home")
            else:
                profile = None
                if hasattr(user, 'staff'):
                    profile = user.staff
                elif hasattr(user, 'faculty'):
                    profile = user.faculty
                elif hasattr(user, 'parents'):
                    profile = user.parents
                elif hasattr(user, 'students'):
                    profile = user.students

                if profile is not None and not profile.is_approved:
                    messages.error(request, "Your account is not approved yet")
                    return redirect("home")
                else:
                    login(request, user)
                    if user.groups.filter(name="Admin").exists():
                        return redirect("register_approval")
                    elif user.groups.filter(name="Staff").exists():
                        return redirect("staff_portal")
                    elif user.groups.filter(name="Faculty").exists():
                        return redirect("faculty_portal")
                    elif user.groups.filter(name="Student").exists():
                        return redirect("student_portal")
                    elif user.groups.filter(name="Parent").exists():
                        return redirect("parent_portal")
                    else:
                        messages.error(request, "User role not defined")
                        return redirect("home")
        
        elif 'register' in request.POST:
            user_type = request.POST.get("user_type")
            registration_form = RegistrationForm(request.POST)

            if user_type == "staff":
                staff_form = StaffForm(request.POST)
                if registration_form.is_valid() and staff_form.is_valid():
                    if User.objects.filter(username=request.POST.get("username")).exists():
                        messages.error(request, "Username already exists. Please choose a different username.")
                        return redirect("home")

                    staff = staff_form.save(commit=False)
                    staff.user = registration_form.save()
                    staff.is_approved = False
                    staff.save()
                    staff_group = Group.objects.get(name="Staff")
                    staff_group.user_set.add(staff.user)
                    return render(request, 'Home/Register_success.html')
            
            elif user_type == "faculty":
                faculty_form = FacultyForm(request.POST)
                if registration_form.is_valid() and faculty_form.is_valid():
                    if User.objects.filter(username=request.POST.get("username")).exists():
                        messages.error(request, "Username already exists. Please choose a different username.")
                        return redirect("home")

                    faculty = faculty_form.save(commit=False)
                    faculty.user = registration_form.save()
                    faculty.is_approved = False
                    faculty.save()
                    faculty_group = Group.objects.get(name="Faculty")
                    faculty_group.user_set.add(faculty.user)
                    return render(request, 'Home/Register_success.html')

            elif user_type == "students":
                students_form = StudentsForm(request.POST)
                if registration_form.is_valid() and students_form.is_valid():
                    if User.objects.filter(username=request.POST.get("username")).exists():
                        messages.error(request, "Username already exists. Please choose a different username.")
                        return redirect("home")

                    students = students_form.save(commit=False)
                    students.user = registration_form.save()
                    students.is_approved = False
                    students.save()
                    student_group = Group.objects.get(name="Student")
                    student_group.user_set.add(students.user)
                    return render(request, 'Home/Register_success.html')

            elif user_type == "parents":
                parents_form = ParentsForm(request.POST)
                if registration_form.is_valid() and parents_form.is_valid():
                    if User.objects.filter(username=request.POST.get("username")).exists():
                        messages.error(request, "Username already exists. Please choose a different username.")
                        return redirect("home")

                    parents = parents_form.save(commit=False)
                    parents.user = registration_form.save()
                    parents.is_approved = False
                    parents.save()
                    parent_group = Group.objects.get(name="Parent")
                    parent_group.user_set.add(parents.user)
                    parents_form.save_m2m()
                    return render(request, 'Home/Register_success.html')

    sections_json = serialize('json', sections, fields=('id', 'name', 'class_linked'))  

    return render(
        request,
        "Home/Home.html",
        {
            "registration_form": registration_form,
            "staff_form": staff_form,
            "faculty_form": faculty_form,
            "students_form": students_form,
            "parents_form": parents_form,
            "sections_json": sections_json,
        },
    )

def about(request):
    return render(request, 'Home/About.html')

def contact(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('name')
            form.save()
            return render(request, 'Home/Contact_success.html', {'name': name})
    else:
        form = ContactForm()
    return render(request, 'Home/Contact.html', {'form': form})

def admission(request):
    if request.method == 'POST':
        form = AdmissionForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('name')
            form.save()
            return render(request, 'Home/Admission_success.html', {'name': name})
    else:
        form = AdmissionForm()
    return render(request, 'Home/Admission.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, "You have been successfully logged out.")
    return redirect("home")


# Admin
@superuser_required
def admission_messages(request):
    admissions = StudentAdmission.objects.all().order_by('-created_at')
    return render(request, 'Admin/admission_messages.html', {'admissions': admissions})

@superuser_required
def contact_messages(request):
    messages = ContactMessage.objects.all().order_by('-created_at')
    return render(request, 'Admin/contact_messages.html', {'messages': messages})

@superuser_required
def register_approvals(request):
    # Get counts of users
    student_count = Students.objects.count()
    parent_count = Parents.objects.count()
    faculty_count = Faculty.objects.count()
    staff_count = Staff.objects.count()

    # Fetch pending and approved users
    pending_users = {
        'staff': Staff.objects.filter(is_approved=False),
        'faculty': Faculty.objects.filter(is_approved=False),
        'students': Students.objects.filter(is_approved=False),
        'parents': Parents.objects.filter(is_approved=False),
    }
    approved_users = {
        'staff': Staff.objects.filter(is_approved=True),
        'faculty': Faculty.objects.filter(is_approved=True),
        'students': Students.objects.filter(is_approved=True),
        'parents': Parents.objects.filter(is_approved=True),
    }

    # Query payments and prepare graph data
    payments = (
        Payment.objects
        .annotate(month=TruncMonth('created_at'))
        .values('month', 'fees_type')
        .annotate(total_amount=Sum('amount'))
        .order_by('month')
    )

    fee_types = ['Tuition_Fees', 'Bus_Fees', 'Mess_Fees']
    chart_data = {fee: [] for fee in fee_types}
    months = []

    for payment in payments:
        month = payment['month'].strftime('%B %Y')
        if month not in months:
            months.append(month)
        fee_key = payment['fees_type'].replace(' ', '_')
        chart_data[fee_key].append(float(payment['total_amount']))  # Convert Decimal to float

    # Ensure all fee types have data for every month
    for fee in fee_types:
        while len(chart_data[fee]) < len(months):
            chart_data[fee].append(0)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            class_id = data.get('class_id')

            if not class_id:
                # If no class_id is provided, default to the first class
                first_class = Class.objects.order_by('id').first()
                if not first_class:
                    return JsonResponse({'error': 'No classes available'}, status=404)
                class_id = first_class.id
            selected_class = Class.objects.get(id=class_id)
            sections = Section.objects.filter(class_linked=selected_class)

            sections_names = []
            months = []
            present_percentages = defaultdict(list)
            absent_percentages = defaultdict(list)

            # Process each section
            for section in sections:
                sections_names.append(section.name)
                
                # Get all attendance records for the section
                attendance_records = StudentAttendance.objects.filter(
                    class_assigned=selected_class, section_assigned=section
                )

                # Group attendance records by month
                month_data = defaultdict(lambda: {'present': 0, 'absent': 0, 'total': 0})

                for record in attendance_records:
                    file = record.file
                    with open(file.path, 'r') as f:
                        csv_reader = csv.DictReader(f)
                        for row in csv_reader:
                            # Extract month and year from date field
                            record_date = record.date
                            month_year = record_date.strftime('%B %Y')  
                            if month_year not in months:
                                months.append(month_year)

                            # Count present and absent students
                            if row['status'] == 'Present':
                                month_data[month_year]['present'] += 1
                            elif row['status'] == 'Absent':
                                month_data[month_year]['absent'] += 1
                            month_data[month_year]['total'] += 1

                # Calculate percentages per month
                for month in months:
                    present_count = month_data[month]['present']
                    absent_count = month_data[month]['absent']
                    total_students = month_data[month]['total']

                    if total_students > 0:
                        present_percentage = (present_count / total_students) * 100
                        absent_percentage = (absent_count / total_students) * 100
                    else:
                        present_percentage = absent_percentage = 0

                    present_percentages[section.name].append(present_percentage)
                    absent_percentages[section.name].append(absent_percentage)

            # Return data as JSON for the graph
            return JsonResponse({
                'sections': sections_names,
                'months': months,
                'present_percentages': dict(present_percentages),
                'absent_percentages': dict(absent_percentages),
                'selected_class_name': selected_class.name,
            })

        except Class.DoesNotExist:
            return JsonResponse({'error': 'Class not found'}, status=404)

    else:
        classes = Class.objects.all()
    # Prepare context data
    context = {
        'pending_users': pending_users,
        'approved_users': approved_users,
        'student_count': student_count,
        'parent_count': parent_count,
        'faculty_count': faculty_count,
        'staff_count': staff_count,
        'chart_data': chart_data,
        'months': json.dumps(months),
        'classes': classes,
    }
    return render(request, 'Admin/register_approvals.html', context)

@superuser_required
def approve_user(request, user_id, user_type):
    user_model = {
        'staff': Staff,
        'faculty': Faculty,
        'students': Students,
        'parents': Parents,
    }.get(user_type)
    user_instance = get_object_or_404(user_model, pk=user_id, is_approved=False)
    user_instance.is_approved = True
    user_instance.save()
    return redirect('register_approval')

@superuser_required
def reject_user(request, user_id, user_type):
    user_model = {
        'staff': Staff,
        'faculty': Faculty,
        'students': Students,
        'parents': Parents,
    }.get(user_type)
    user_instance = get_object_or_404(user_model, pk=user_id, is_approved=False)
    user_instance.delete()
    return redirect('register_approval')

@superuser_required
def classes(request):
    classes = Class.objects.all()
    sorted_classes = sorted(classes, key=lambda x: int(x.name))

    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            class_name = form.cleaned_data['name']
            if Class.objects.filter(name=class_name).exists():
                messages.error(request, f'Class "{class_name}" already exists.')
            else:
                form.save()
                return redirect('classes')
    else:
        form = ClassForm()
    return render(request, 'Admin/classes.html', {'form': form, 'sorted_classes': sorted_classes})

@superuser_required
def update_class(request, class_id):
    class_instance = get_object_or_404(Class, id=class_id)
    if request.method == 'POST':
        form = ClassForm(request.POST, instance=class_instance)
        if form.is_valid():
            form.save()
            return redirect('classes') 
    else:
        form = ClassForm(instance=class_instance)
    return render(request, 'Admin/classes.html', {'form': form, 'class_instance': class_instance})

@superuser_required
def delete_class(request, class_id):
    class_instance = get_object_or_404(Class, id=class_id)
    if request.method == 'POST':
        class_instance.delete()
        return redirect('classes')
    return render(request, 'Admin/classes.html', {'class_instance': class_instance})

@superuser_required
def sections(request):
    classes = Class.objects.all()
    classes_with_sections = Class.objects.filter(section__isnull=False).distinct()
    sorted_classes = sorted(classes_with_sections, key=lambda x: int(x.name))
    sections_by_class = {}
    for class_obj in sorted_classes:
        sections = Section.objects.filter(class_linked=class_obj).order_by('name')
        sections_by_class[class_obj] = sections
    
    # create section
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            class_id = form.cleaned_data['class_linked']
            section_name = form.cleaned_data['name']
            if Section.objects.filter(class_linked=class_id, name=section_name).exists():
                messages.error(request, f'Section "{section_name}" for class "{class_id.name}" already exists.')
            else:
                form.save()
                return redirect('sections')
    else:
        form = SectionForm()
    return render(request, 'Admin/sections.html', {'form': form,'classes': classes,'sections_by_class': sections_by_class})

@superuser_required
def update_sections(request):
    if request.method == 'POST':
        sections_data = request.POST.getlist('sections[]')
        class_id = request.POST.get('class_linked')
        try:
            sections_to_update = Section.objects.filter(
                class_linked_id=class_id
            )
            # Ensure the number of new names matches the number of existing sections
            if len(sections_data) != sections_to_update.count():
                return HttpResponseBadRequest('Number of sections mismatch.')
            for index, section in enumerate(sections_to_update):
                section.name = sections_data[index]
                section.save()
            return redirect('sections')  
        except Exception as e:
            return HttpResponseBadRequest('Error updating sections.')

    else:
        return redirect('sections')

@superuser_required
def delete_sections(request):
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        sections_to_delete = request.POST.getlist('sections_to_delete[]')
        if not class_id:
            return HttpResponseBadRequest('Class ID is required.')
        if not sections_to_delete:
            return HttpResponseBadRequest('No sections selected for deletion.')
        try:
            Section.objects.filter(id__in=sections_to_delete).delete()
            return redirect('sections')  
        except Exception as e:
            return HttpResponseBadRequest('Error deleting sections.')
    else:
        return redirect('sections') 
    
@superuser_required
def subjects(request):
    subjects_by_class_section = {}
    subjects = Subject.objects.select_related('class_linked', 'section_linked').all()
    for subject in subjects:
        class_linked = subject.class_linked
        section_linked = subject.section_linked
        if class_linked is None or section_linked is None:
            continue

        class_section_key = (class_linked, section_linked)
        if class_section_key not in subjects_by_class_section:
            subjects_by_class_section[class_section_key] = []
        subjects_by_class_section[class_section_key].append(subject)

    sorted_keys = sorted(subjects_by_class_section.keys(), key=lambda x: (int(x[0].name), x[1].name))
    sorted_subjects_by_class_section = {key: subjects_by_class_section[key] for key in sorted_keys}

    # create subject
    if request.method == 'POST':
        formset = SubjectFormSet(request.POST)
        if formset.is_valid():
            class_id = request.POST.get('class_linked')
            section_id = request.POST.get('section_linked')

            # Get valid forms only
            subjects_data = [form.cleaned_data['subjectname'] for form in formset if 'subjectname' in form.cleaned_data]
            # subjects_data = [form.cleaned_data['subjectname'] for form in formset]
            existing_subjects = Subject.objects.filter(class_linked_id=class_id, section_linked_id=section_id, subjectname__in=subjects_data)
            if existing_subjects.exists():
                class_name = Class.objects.get(id=class_id).name
                section_name = Section.objects.get(id=section_id).name
                existing_names = ', '.join(existing_subjects.values_list('subjectname', flat=True))
                messages.error(request, f'The following subjects already exist for class "{class_name}"-section "{section_name}" combination: {existing_names}')
            else:
                for form in formset:
                    subject = form.save(commit=False)
                    subject.class_linked_id = class_id
                    subject.section_linked_id = section_id
                    subject.save()
                return redirect('subjects')
    else:
        formset = SubjectFormSet()
    classes = Class.objects.all()
    sections = Section.objects.all()
    sections_json = serialize('json', sections, fields=('id', 'name', 'class_linked'))    
    return render(request, 'Admin/subjects.html', {'formset': formset, 'classes': classes,'sections':sections, 'sections_json': sections_json, 'subjects':subjects,'subjects_by_class_section': sorted_subjects_by_class_section})

@superuser_required
def update_subjects(request):
    if request.method == 'POST':
        subjects_data = request.POST.getlist('subjects[]')
        class_id = request.POST.get('class_linked')
        section_id = request.POST.get('section_linked')
        try:
            subjects_to_update = Subject.objects.filter(
                class_linked_id=class_id,
                section_linked_id=section_id
            )
            # Ensure the number of new names matches the number of existing subjects
            if len(subjects_data) != subjects_to_update.count():
                return HttpResponseBadRequest('Number of subjects mismatch.')
            for index, subject in enumerate(subjects_to_update):
                subject.subjectname = subjects_data[index]
                subject.save()
            return redirect('subjects')  
        except Exception as e:
            return HttpResponseBadRequest('Error updating subjects.')

    else:
        return redirect('subjects')

@superuser_required    
def delete_subjects(request):
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        section_id = request.POST.get('section_id')
        try:
            subjects_to_delete = request.POST.getlist('subjects_to_delete[]')
            Subject.objects.filter(
                class_linked_id=class_id,
                section_linked_id=section_id,
                id__in=subjects_to_delete
            ).delete()
            return redirect('subjects')  # Redirect to wherever you need after deleting subjects
        except Exception as e:
            return HttpResponseBadRequest('Error deleting subjects.')

    return redirect('subjects')     
    
@superuser_required
def exam(request):
    # upload exam timetable
    form = ExamForm()
    classes = Class.objects.all()
    sections = Section.objects.all()
    exam_types = [exam_type[0] for exam_type in Exam.EXAM_TYPES]
    error_message = None

    if request.method == 'POST':
        form = ExamForm(request.POST, request.FILES)
        if form.is_valid():
            class_id = request.POST.get('class_linked')
            section_id = request.POST.get('section_linked')
            exam_type = request.POST.get('exam_type')
            unit_test_number = request.POST.get('unit_test_number')

            # Check if exam timetable for the same class-section and exam type already exist
            existing_exam = Exam.objects.filter(
                class_linked_id=class_id,
                section_linked_id=section_id,
                exam_type=exam_type,
            )
            if exam_type == 'Unit Test':
                existing_exam = existing_exam.filter(unit_test_number=unit_test_number)

            if existing_exam.exists():
                error_message = "Exam Timetable for this class-section and exam type have already been uploaded."
            else:
                student_exam = form.save(commit=False)
                student_exam.class_linked_id = class_id
                student_exam.section_linked_id = section_id
                student_exam.exam_type = exam_type
                if exam_type == 'Unit Test':
                    student_exam.unit_test_number = unit_test_number
                else:
                    student_exam.unit_test_number = None
                student_exam.save()
                return redirect('exam')
        else:
            logger.debug("Form is not valid")
            logger.debug("Form errors: %s", form.errors)

    
    sections_json = serialize('json', sections, fields=('id', 'name', 'class_linked'))

    # view exam timetable
    student_exam = Exam.objects.filter()
    exam_type_dict = {}
    for exam_type in exam_types:
        if exam_type == 'Unit Test':
            unit_test_dict = {}
            unit_test_numbers = student_exam.filter(exam_type='Unit Test').values_list('unit_test_number', flat=True).distinct()
            for unit_test_number in unit_test_numbers:
                unit_test_dict[unit_test_number] = student_exam.filter(exam_type='Unit Test', unit_test_number=unit_test_number)
            exam_type_dict['Unit Test'] = unit_test_dict
        else:
            exam_type_dict[exam_type] = student_exam.filter(exam_type=exam_type)

    return render(request, 'Admin/exam.html', {
        'form':form,
        'classes': classes,
        'sections': sections,
        'exam_types':exam_types,
        'sections_json':sections_json ,
        'exam_type_dict': exam_type_dict,
        'error_message': error_message
    })

@superuser_required
def departments(request):
    departments = Department.objects.all()
    faculties = Faculty.objects.all()

    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department_name = form.cleaned_data['name']
            selected_faculty_ids = request.POST.getlist('faculty')
            if Department.objects.filter(name=department_name).exists():
                messages.error(request, f'Department with name "{department_name}" already exists.')
            else:
                invalid_faculties = []
                for faculty_id in selected_faculty_ids:
                    try:
                        faculty = Faculty.objects.get(faculty_id=faculty_id)
                        if faculty.departments.exists():
                            invalid_faculties.append(faculty.name)
                    except Faculty.DoesNotExist:
                        invalid_faculties.append(faculty_id)
                
                if invalid_faculties:
                    invalid_faculty_names = ', '.join(invalid_faculties)
                    messages.error(request, f'The following faculty members are invalid or already assigned to another department: {invalid_faculty_names}')           
                else:
                    form.save()
                    return redirect('departments') 
    else:
        form = DepartmentForm()
    return render(request, 'Admin/departments.html', {'form': form, 'departments': departments, 'faculties':faculties})

@superuser_required
def edit_department(request, department_id):
    department = get_object_or_404(Department, id=department_id)
    
    if request.method == 'POST':
        new_name = request.POST.get('name')
        original_name = request.POST.get('original_name')
        faculty_ids_to_remove = request.POST.getlist('faculty_to_remove')
        faculty_ids_to_add = request.POST.getlist('faculty_to_add')
        
        # Update department name if changed
        if new_name != original_name:
            if Department.objects.filter(name=new_name).exists():
                messages.error(request, f'A department with the name "{new_name}" already exists.')
                return redirect('departments')
            department.name = new_name
            department.save()
        
        # Check if faculty members to add are not already assigned to another department
        invalid_faculties = []
        for faculty_id in faculty_ids_to_add:
            try:
                faculty = Faculty.objects.get(faculty_id=faculty_id)
                if faculty.departments.exists():
                    invalid_faculties.append(faculty.name)
            except Faculty.DoesNotExist:
                invalid_faculties.append(faculty_id)
        
        if invalid_faculties:
            invalid_faculty_names = ', '.join(invalid_faculties)
            messages.error(request, f'The following faculty members are already assigned to another department: {invalid_faculty_names}')
            return redirect('departments')
        
        # Process adding new faculty
        for faculty_id in faculty_ids_to_add:
            faculty = get_object_or_404(Faculty, faculty_id=int(faculty_id))
            department.faculty.add(faculty)
        
        # Remove selected faculties from the department
        for faculty_id in faculty_ids_to_remove:
            faculty = get_object_or_404(Faculty, faculty_id=int(faculty_id))
            department.faculty.remove(faculty)
        return redirect('departments') 
    faculties = Faculty.objects.all()
    return render(request, 'Admin/departments.html', {'department': department, 'faculties':faculties})

@superuser_required
def academic_session(request):
    academics = AcademicSession.objects.all()

    if request.method == 'POST':
        form = AcademicSessionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('academic_session') 
    else:
        form = AcademicSessionForm()
    return render(request, 'Admin/academic_session.html', {'form': form, 'academics':academics})

@superuser_required
def employee_category(request):
    faculties = Faculty.objects.all()
    staffs = Staff.objects.all()
    return render(request, 'Admin/employee_category.html', {'faculties': faculties, 'staffs': staffs})

@superuser_required
def leave_approval(request):
    pending_requests = LeaveRequest.objects.filter(status='Pending')
    approved_requests = LeaveRequest.objects.filter(status='Approved')
    declined_requests = LeaveRequest.objects.filter(status='Declined')

    staff = StaffLeave.objects.all()
    faculty = FacultyLeave.objects.all()
    
    staff_error = None
    faculty_error = None

    if request.method == 'POST':
        if 'staff_submit' in request.POST:
            staff_form = StaffLeaveForm(request.POST, request.FILES)
            faculty_form = FacultyLeaveForm()
            if staff_form.is_valid():
                leave_month = staff_form.cleaned_data['month']
                if StaffLeave.objects.filter(month=leave_month).exists():
                    staff_error = f'Staff leave for {leave_month} already exists.'
                else:
                    staff_form.save()
                    return redirect('leave_approval')
        elif 'faculty_submit' in request.POST:
            faculty_form = FacultyLeaveForm(request.POST, request.FILES)
            staff_form = StaffLeaveForm()
            if faculty_form.is_valid():
                leave_month = faculty_form.cleaned_data['month']
                if FacultyLeave.objects.filter(month=leave_month).exists():
                    faculty_error = f'Faculty leave for {leave_month} already exists.'
                else:
                    faculty_form.save()
                    return redirect('leave_approval')
    else:
        staff_form = StaffLeaveForm()
        faculty_form = FacultyLeaveForm()

    return render(request, 'Admin/leave_approval.html',
                   {'pending_requests': pending_requests,
                    'approved_requests': approved_requests,
                    'declined_requests': declined_requests,
                    'staff_form':staff_form,
                    'faculty_form':faculty_form,
                    'staff':staff,
                    'faculty':faculty,
                    'staff_error': staff_error,
                    'faculty_error': faculty_error
                    })


@superuser_required
def update_leave_approval(request, pk, status):
    leave_request = get_object_or_404(LeaveRequest, pk=pk)
    leave_request.status = status
    leave_request.save()
    return redirect('leave_approval')

@superuser_required
def fees(request):
    bus = BusFees.objects.all()
    tuition = TuitionFees.objects.all()
    mess = MessFees.objects.all()
    
    bus_error = None
    tuition_error = None
    mess_error = None

    if request.method == 'POST':
        if 'bus_submit' in request.POST:
            bus_form = BusFeesForm(request.POST, request.FILES)
            tuition_form = TuitionFeesForm()
            mess_form = MessFeesForm()
            if bus_form.is_valid():
                location = bus_form.cleaned_data['location']
                if BusFees.objects.filter(location=location).exists():
                    bus_error = f'Bus fees for location {location} already exist.'
                else:
                    bus_form.save()
                    return redirect('fees')
        elif 'tuition_submit' in request.POST:
            tuition_form = TuitionFeesForm(request.POST, request.FILES)
            bus_form = BusFeesForm()
            mess_form = MessFeesForm()
            if tuition_form.is_valid():
                class_linked = tuition_form.cleaned_data['class_linked']
                if TuitionFees.objects.filter(class_linked= class_linked).exists():
                    tuition_error = f'Tuition fees for class {class_linked} already exist.'
                else:
                    tuition_form.save()
                    return redirect('fees')
        elif 'mess_submit' in request.POST:
            mess_form = MessFeesForm(request.POST, request.FILES)
            bus_form = BusFeesForm()
            tuition_form = TuitionFeesForm()
            if mess_form.is_valid():
                if MessFees.objects.exists():
                    mess_error = 'Mess fees already exist.'
                else:
                    mess_form.save()
                    return redirect('fees')
    else:
        bus_form = BusFeesForm()
        tuition_form = TuitionFeesForm()
        mess_form = MessFeesForm()

    return render(request, 'Admin/fees.html', {
        'bus_form': bus_form,
        'tuition_form': tuition_form,
        'mess_form': mess_form,
        'bus': bus,
        'tuition': tuition,
        'mess': mess,
        'bus_error': bus_error,
        'tuition_error': tuition_error,
        'mess_error': mess_error
    })

@superuser_required
def fees_payment(request):
    classes = Class.objects.all()
    sections = Section.objects.all()
    students = Students.objects.all()
    payment_types = [payment_type[0] for payment_type in Payment.PAYMENT_CHOICES]
    fees_types = [fees_type[0] for fees_type in Payment.FEES_CHOICES]

    if request.method == 'POST':
        class_id = request.POST.get('class_linked')
        section_id = request.POST.get('section_linked')
        student_id = request.POST.get('student')
        receipt_number = request.POST.get('receipt_number')

        for fees_type in fees_types:
            payment_type = request.POST.get(f'{fees_type}_payment_type')
            amount = request.POST.get(f'{fees_type}_amount')

            if payment_type and amount: 
                Payment.objects.create(
                    student_id=student_id,
                    class_linked_id=class_id,
                    section_linked_id=section_id,
                    payment_type=payment_type,
                    fees_type=fees_type,
                    amount=amount,
                    receipt_number=receipt_number
                )
        return redirect('fees_payment')
    
    sections_json = serialize('json', sections, fields=('id', 'name', 'class_linked'))
    students_json = serialize('json', students, fields=('id', 'name', 'section_linked'))

    # view fees
    payments_by_class_section_student = {}
    for cls in classes:
        payments_by_class_section_student[cls] = {}
        for section in cls.section_set.all():
            payments_by_class_section_student[cls][section] = []
            for student in section.students_set.all():
                try:
                    tuition_fee_obj = TuitionFees.objects.get(class_linked=cls)
                    tuition_fee_total = tuition_fee_obj.amount
                except TuitionFees.DoesNotExist:
                    tuition_fee_total = 0
                mess_fee_total = MessFees.objects.first().amount 
                try:
                    bus_fee_obj = BusFees.objects.get(location=student.location)
                    bus_fee_total = bus_fee_obj.amount
                except BusFees.DoesNotExist:
                    bus_fee_total = 0
                payments = Payment.objects.filter(
                    student=student,
                    class_linked=cls,
                    section_linked=section
                )
                student_payments = {'student_name': student.name, 'Bus_Fees': 0, 'Mess_Fees': 0, 'Tuition_Fees': 0}
                for payment in payments:
                    if payment.fees_type == 'Bus Fees':
                        student_payments['Bus_Fees'] += payment.amount
                    elif payment.fees_type == 'Mess Fees':
                        student_payments['Mess_Fees'] += payment.amount
                    elif payment.fees_type == 'Tuition Fees':
                        student_payments['Tuition_Fees'] += payment.amount

                # Calculate pending amounts
                tuition_pending = max(0, tuition_fee_total - student_payments['Tuition_Fees'])
                mess_pending = max(0, mess_fee_total - student_payments['Mess_Fees'])
                bus_pending = max(0, bus_fee_total - student_payments['Bus_Fees'])

                student_payments['Bus_Pending'] = bus_pending
                student_payments['Mess_Pending'] = mess_pending
                student_payments['Tuition_Pending'] = tuition_pending

                payments_by_class_section_student[cls][section].append(student_payments)

    return render(request, 'Admin/fees_payment.html', {
        'classes': classes,
        'sections': sections,
        'students': students,
        'payment_types': payment_types,
        'fees_types': fees_types,
        'sections_json': sections_json,
        'students_json': students_json,
        'payments_by_class_section_student': payments_by_class_section_student,
    })

@superuser_required
def salary_payment(request):
    staff_payslips = StaffSalary.objects.all()
    faculty_payslips = FacultySalary.objects.all()
    
    staff_error = None
    faculty_error = None

    if request.method == 'POST':
        if 'staff_submit' in request.POST:
            staff_form = StaffSalaryForm(request.POST, request.FILES)
            faculty_form = FacultySalaryForm()  # Ensure faculty form is empty
            if staff_form.is_valid():
                staff_member = staff_form.cleaned_data['staff']
                month = staff_form.cleaned_data['month']
                year = staff_form.cleaned_data['year']

                if StaffSalary.objects.filter(staff=staff_member, month=month, year=year).exists():
                    staff_error = f'Salary for {staff_member} for {month}/{year} already exists.'
                else:
                    staff_form.save()
                    return redirect('salary_payment')
        elif 'faculty_submit' in request.POST:
            faculty_form = FacultySalaryForm(request.POST, request.FILES)
            staff_form = StaffSalaryForm()  # Ensure staff form is empty
            if faculty_form.is_valid():
                faculty_member = faculty_form.cleaned_data['faculty']
                month = faculty_form.cleaned_data['month']
                year = faculty_form.cleaned_data['year']

                if FacultySalary.objects.filter(faculty=faculty_member, month=month, year=year).exists():
                    faculty_error = f'Salary for {faculty_member} for {month}/{year} already exists.'
                else:
                    faculty_form.save()
                    return redirect('salary_payment')
    else:
        staff_form = StaffSalaryForm()
        faculty_form = FacultySalaryForm()  
    return render(request, 'Admin/salary_payment.html', {'staff_form': staff_form, 'faculty_form':faculty_form, 'staff_payslips':staff_payslips,'faculty_payslips':faculty_payslips,
        'staff_error': staff_error,
        'faculty_error': faculty_error})

@superuser_required
def busroute(request):
    bus_routes = BusRoute.objects.prefetch_related('pickup_locations', 'dropoff_locations').all()
    
    if request.method == 'POST':
        form = BusRouteForm(request.POST)
        if form.is_valid():
            bus_route = form.save(commit=False)
            bus_route.route_name = request.POST.get('route_name')
            bus_route.bus_number = request.POST.get('bus_number')
            bus_route.driver_name = request.POST.get('driver_name')
            bus_route.driver_number = request.POST.get('driver_number')
            bus_route.save()

            additional_pickup_locations = request.POST.getlist('pickup_location')
            additional_pickup_times = request.POST.getlist('pickup_time')
            additional_dropoff_locations = request.POST.getlist('dropoff_location')
            additional_dropoff_times = request.POST.getlist('dropoff_time')

            for location, time in zip(additional_pickup_locations, additional_pickup_times):
                PickupLocation.objects.create(bus_route=bus_route, location=location, time=time)

            for location, time in zip(additional_dropoff_locations, additional_dropoff_times):
                DropoffLocation.objects.create(bus_route=bus_route, location=location, time=time)

            return redirect('busroute') 
    else:
        form = BusRouteForm()
    
    return render(request, 'Admin/busroute.html', {'form': form, 'bus_routes': bus_routes})

#send sms
@superuser_required
def send_sms(request):
    if request.method == 'POST':
        message_content = request.POST.get('message')
        if message_content:
            parents = Parents.objects.all()
            for parent in parents:
                phone_number = parent.contact_number
                full_phone_number = '+91' + phone_number
                send_sms_to_phone(full_phone_number, message_content)
            messages.success(request, 'Message sent to all parents successfully.')
            return redirect('send_sms')
    return render(request, 'Admin/send_sms.html')

def send_sms_to_phone(full_phone_number, message):
    client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
    client.messages.create(
        body=message,
        from_=settings.TWILIO_PHONE_NUMBER,
        to=full_phone_number
    )

@superuser_required
def attendance(request):
    today = timezone.now().date()
    form = FacultyAttendanceForm()
    staff_form = StaffAttendanceForm()
    faculty_attendance_exists = FacultyAttendance.objects.filter(
        date=today
    ).exists()

    staff_attendance_exists = StaffAttendance.objects.filter(
        date=today
    ).exists()

    faculty_messages = []
    staff_messages = []

    if faculty_attendance_exists:
        faculty_messages.append('Attendance for today has already been recorded for faculty.')
    if staff_attendance_exists:
        staff_messages.append('Attendance for today has already been recorded for staff.')
    if request.method == 'POST':
        if not faculty_attendance_exists and 'submit_faculty' in request.POST:
            form = FacultyAttendanceForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                return redirect('attendance')
        elif not staff_attendance_exists and 'submit_staff' in request.POST:
            staff_form = StaffAttendanceForm(request.POST, request.FILES)
            if staff_form.is_valid():
                staff_form.save()
                return redirect('attendance')
    # view attendance
    faculty_attendance_records = FacultyAttendance.objects.all().order_by('-date')
    staff_attendance_records = StaffAttendance.objects.all().order_by('-date')  

    return render(request, 'Admin/attendance.html', {'form': form, 'staff_form': staff_form,
        'faculty_messages': faculty_messages,
        'staff_messages': staff_messages,
        'faculty_attendance_records': faculty_attendance_records,
        'staff_attendance_records': staff_attendance_records})

@superuser_required
def timetable(request):
    student_form = StudentTimetableForm() 
    faculty_form = FacultyTimetableForm()
    staff_form = StaffTimetableForm() 
    if request.method == 'POST':
        if 'submit_student' in request.POST:
            student_form = StudentTimetableForm(request.POST, request.FILES)
            if student_form.is_valid():
                class_id = request.POST.get('class_linked')
                section_id = request.POST.get('section_linked')
                # Check if timetable already exists for this class and section combination
                if not StudentTimetable.objects.filter(class_linked_id=class_id, section_linked_id=section_id).exists():
                    student_timetable = student_form.save(commit=False)
                    student_timetable.class_linked_id = class_id
                    student_timetable.section_linked_id = section_id
                    student_timetable.save()
                    return redirect('timetable')
                else:
                    student_form.add_error(None, 'Timetable already exists for this class and section combination.')
        elif 'submit_faculty' in request.POST:
            faculty_form = FacultyTimetableForm(request.POST, request.FILES)
            if faculty_form.is_valid():
                # Check if timetable already exists for this faculty
                if not FacultyTimetable.objects.filter(faculty=faculty_form.cleaned_data['faculty']).exists():
                    faculty_form.save()
                    return redirect('timetable')
                else:
                    faculty_form.add_error(None, 'Timetable already exists for this faculty.')
        elif 'submit_staff' in request.POST:
            staff_form = StaffTimetableForm(request.POST, request.FILES)
            if staff_form.is_valid():
                # Check if timetable already exists for this staff
                if not StaffTimetable.objects.filter(staff=staff_form.cleaned_data['staff']).exists():
                    staff_form.save()
                    return redirect('timetable')
                else:
                    staff_form.add_error(None, 'Timetable already exists for this staff.')  
    # view student timetable
    student_timetables = StudentTimetable.objects.all()
    # view faaculty timetable
    faculty_timetables = FacultyTimetable.objects.all()
    # view staff timetable
    staff_timetables = StaffTimetable.objects.all()
    classes = Class.objects.all()
    sections = Section.objects.all()    
    sections_json = serialize('json', sections, fields=('id', 'name', 'class_linked'))

    context = {'faculty_form':faculty_form,
               'faculty_timetables':faculty_timetables,
               'staff_form':staff_form,
               'staff_timetables':staff_timetables,
               'student_form':student_form,
               'student_timetables':student_timetables,
               'classes':classes, 
               'sections':sections, 
               'sections_json':sections_json,}
    return render(request, 'Admin/timetable.html', context )
    
    

#Staff
@staff_required
def staff_portal(request):
    student_count = Students.objects.count()
    parent_count = Parents.objects.count()
    faculty_count = Faculty.objects.count()
    staff_count = Staff.objects.count()
    try:
        staff = request.user.staff
        leave_requests = LeaveRequest.objects.filter(user=request.user, notified=False, status_changed__isnull=False)
        if leave_requests.exists():
            messages.info(request, "Your leave request status has been changed.")
            leave_requests.update(notified=True)
        return render(request, 'Staff/staff_portal.html', {'staff': staff,
                'student_count': student_count,
                'parent_count': parent_count,
                'faculty_count': faculty_count,
                'staff_count': staff_count,})
    except Staff.DoesNotExist:
        messages.error(request, "Your registration has been rejected.")
        return redirect('home') 

@staff_required
def staff_attendance(request):
    staff = request.user.staff
    attendance_records = []
    files = StaffAttendance.objects.all().order_by('date')
    for attendance_file in files:
        if attendance_file.file.name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(attendance_file.file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):  
                    if row[0] == staff.name:
                        record = {
                            'date': attendance_file.date,
                            'status': row[1]
                        }
                        attendance_records.append(record)
            except InvalidFileException:
                continue  
            except Exception as e:
                continue
        elif attendance_file.file.name.endswith('.csv'):
            try:
                file_data = attendance_file.file.read().decode('utf-8').splitlines()
                reader = csv.reader(file_data)
                next(reader)  # Skip the header row
                for row in reader:
                    if row[0] == staff.name:
                        record = {
                            'date': attendance_file.date,
                            'status': row[1]
                        }
                        attendance_records.append(record)
            except Exception as e:
                continue
            
    from collections import defaultdict
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'total_days': 0, 'days_absent': 0}))
    for record in attendance_records:
        year = record['date'].year
        month = record['date'].month
        monthly_data[year][month]['total_days'] += 1
        if record['status'] == 'Absent':
            monthly_data[year][month]['days_absent'] += 1
    attendance_by_year = {}
    for year, months in monthly_data.items():
        attendance_by_year[year] = []
        for month, data in months.items():
            total_days = data['total_days']
            days_absent = data['days_absent']
            days_present = total_days - days_absent
            attendance_percentage = (days_present / total_days) * 100 if total_days > 0 else 0
            attendance_by_year[year].append({
                'month': calendar.month_name[month],
                'total_days': total_days,
                'days_absent': days_absent,
                'days_present': days_present,
                'attendance_percentage': attendance_percentage
            })

    context = {"staff": staff,'attendance_by_year':attendance_by_year}
    return render(request, 'Staff/attendance.html',context)

@staff_required
def staff_timetable(request):
    staff = request.user.staff
    timetable = StaffTimetable.objects.filter(staff=staff).last()
    timetable_data = []
    if timetable:
        file_path = timetable.file.path
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    timetable_data.append(row)
        elif file_path.endswith('.xlsx'):
            timetable_df = pd.read_excel(file_path)
            timetable_data = timetable_df.to_dict(orient='records')
        else:
            print("Unsupported file format for timetable:", file_path)
            timetable_data = None
    else:
        print("No Timetable instance found for faculty:", staff.staff_id)
        timetable_data = None
    return render(request, 'Staff/timetable.html', {'staff': staff,'timetable_data':timetable_data})

@staff_required
def staff_leave_request(request):
    staff = request.user.staff
    leave_requests = LeaveRequest.objects.filter(user=request.user)
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.user = request.user
            leave_request.status = 'Pending'  # Set status to 'Pending'
            leave_request.save()
            return redirect('staff_leave_request')  
    else:
        form = LeaveRequestForm()
    
    user = request.user
    staff_leaves = StaffLeave.objects.all()
    months_order = list(calendar.month_name)[1:] 
    valid_staff_leaves = [leave for leave in staff_leaves if leave.month in months_order]
    staff_leaves = sorted(valid_staff_leaves, key=lambda leave: months_order.index(leave.month))
    leave_data = []
    for staff_leave in staff_leaves:
        try:
            month_num = datetime.strptime(staff_leave.month, '%B').month
        except ValueError as e:
            continue
        leaves_requests = LeaveRequest.objects.filter(
            user=user,
            start_date__month=month_num,
            status='Approved'
        )
        leaves_taken = sum((request.end_date - request.start_date).days + 1 for request in leaves_requests)
        remaining_leaves = staff_leave.staff_leaves - leaves_taken

        leave_data.append({
            'month': staff_leave.month,
            'allocated_leaves': staff_leave.staff_leaves,
            'leaves_taken': leaves_taken,
            'remaining_leaves': remaining_leaves
        })

        
    return render(request, 'Staff/leave_request.html',
                   {"staff": staff,
                   'form': form,
                   'leave_requests': leave_requests,
                   'leave_data':leave_data})

@staff_required
def staff_payslip(request):
    staff = request.user.staff
    payslips = StaffSalary.objects.filter(staff=staff)
    return render(request, 'Staff/salary.html', {'staff':staff,'payslips': payslips})

@staff_required
def staff_bus_details(request):
    staff = Staff.objects.get(user=request.user) 
    staff_location = staff.location  
    matching_routes = BusRoute.objects.filter(pickup_locations__location=staff_location)
    pickup_data = []
    dropoff_data = []
    # initialize values
    driver_name = None
    driver_number = None
    bus_number = None
    for route in matching_routes:
        driver_name = route.driver_name
        driver_number = route.driver_number
        bus_number = route.bus_number
        pickup_locations = PickupLocation.objects.filter(bus_route=route)
        dropoff_locations = DropoffLocation.objects.filter(bus_route=route)
        for pickup_location in pickup_locations:
            pickup_data.append({
                'pickup_location': pickup_location.location,
                'pickup_time': pickup_location.time,
                'driver_name': pickup_location.bus_route.driver_name,
                'bus_number': pickup_location.bus_route.bus_number,
                'driver_number': pickup_location.bus_route.driver_number
            })
        for dropoff_location in dropoff_locations:
            dropoff_data.append({
                'dropoff_location': dropoff_location.location,
                'dropoff_time': dropoff_location.time
            })

    context = {
        'staff': staff,
        'pickup_data': pickup_data,
        'dropoff_data': dropoff_data,
        'driver_name': driver_name,
        'driver_number': driver_number,
        'bus_number': bus_number
    }
    return render(request, 'Staff/staff_bus.html', context)

#Faculty
@faculty_required
def faculty_portal(request):
    student_count = Students.objects.count()
    parent_count = Parents.objects.count()
    faculty_count = Faculty.objects.count()
    staff_count = Staff.objects.count()
    try:
        faculty = request.user.faculty
        # Leave Request Notification
        leave_requests = LeaveRequest.objects.filter(user=request.user, notified=False, status_changed__isnull=False)
        if leave_requests.exists():
            messages.info(request, "Your leave request status has been changed.")
            leave_requests.update(notified=True)
        # fetch class-section/subject
        faculty_timetable = FacultyTimetable.objects.filter(faculty=faculty).last()

        if not faculty_timetable:
            context = {
                'faculty': faculty,
                'subjects_classes_sections': [],
                'message': 'No timetable found for the logged-in faculty.',
                'student_count': student_count,
                'parent_count': parent_count,
                'faculty_count': faculty_count,
                'staff_count': staff_count,
            }
            return render(request, 'Faculty/faculty_portal.html', context)
        file_path = faculty_timetable.file.path
        subjects_classes_sections = set()
        period_fields = []

        if file_path.endswith('.csv'):
            with open(file_path, newline='') as csvfile:
                reader = csv.DictReader(csvfile)
                period_fields = reader.fieldnames[1:]  # Skip the first column (assumed to be the day)
                for row in reader:
                    for period_field in period_fields:
                        period_value = row.get(period_field)
                        if period_value and period_value.strip() != '----':
                            if '/' in period_value:
                                subject, class_section = period_value.split('/')
                                subjects_classes_sections.add((subject.strip(), class_section.strip()))
        elif file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]  # Assumes the first row is the header
            period_fields = headers[1:]  # Skip the first column (assumed to be the day)
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                for period_field in period_fields:
                    period_value = row_dict.get(period_field)
                    if period_value and period_value.strip() != '----':
                        if '/' in period_value:
                            subject, class_section = period_value.split('/')
                            subjects_classes_sections.add((subject.strip(), class_section.strip()))

        context = {
            'faculty': faculty,
            'subjects_classes_sections': sorted(subjects_classes_sections),
            'student_count': student_count,
            'parent_count': parent_count,
            'faculty_count': faculty_count,
            'staff_count': staff_count,
        }
        return render(request, 'Faculty/faculty_portal.html', context)
    except Faculty.DoesNotExist:
        messages.error(request, "Your registration has been rejected.")
        return redirect('home') 

@faculty_required
def faculty_attendance(request):
    faculty = request.user.faculty
    attendance_records = []
    files = FacultyAttendance.objects.all().order_by('date')
    for attendance_file in files:
        if attendance_file.file.name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(attendance_file.file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):  
                    if row[0] == faculty.name:
                        record = {
                            'date': attendance_file.date,
                            'status': row[1]
                        }
                        attendance_records.append(record)
            except InvalidFileException:
                continue  
            except Exception as e:
                continue
        elif attendance_file.file.name.endswith('.csv'):
            try:
                file_data = attendance_file.file.read().decode('utf-8').splitlines()
                reader = csv.reader(file_data)
                next(reader)  # Skip the header row
                for row in reader:
                    if row[0] == faculty.name:
                        record = {
                            'date': attendance_file.date,
                            'status': row[1]
                        }
                        attendance_records.append(record)
            except Exception as e:
                continue
            
    from collections import defaultdict
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'total_days': 0, 'days_absent': 0}))
    for record in attendance_records:
        year = record['date'].year
        month = record['date'].month
        monthly_data[year][month]['total_days'] += 1
        if record['status'] == 'Absent':
            monthly_data[year][month]['days_absent'] += 1
    attendance_by_year = {}
    for year, months in monthly_data.items():
        attendance_by_year[year] = []
        for month, data in months.items():
            total_days = data['total_days']
            days_absent = data['days_absent']
            days_present = total_days - days_absent
            attendance_percentage = (days_present / total_days) * 100 if total_days > 0 else 0
            attendance_by_year[year].append({
                'month': calendar.month_name[month],
                'total_days': total_days,
                'days_absent': days_absent,
                'days_present': days_present,
                'attendance_percentage': attendance_percentage
            })

    # upload attendance
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        return redirect('faculty_portal')
    
    form = StudentAttendanceForm()
    today = timezone.now().date()
    attendance_exists = StudentAttendance.objects.filter(
        faculty=faculty,
        date=today
    ).exists()
    if attendance_exists:
        messages.warning(request, 'Attendance for today has already been recorded for your class and section.')
    elif request.method == 'POST':
        form = StudentAttendanceForm(request.POST, request.FILES)
        if form.is_valid():
            attendance_instance = form.save(commit=False)
            attendance_instance.faculty = faculty
            attendance_instance.class_assigned = faculty.class_assigned
            attendance_instance.section_assigned = faculty.section_assigned
            attendance_instance.save()
            return redirect('faculty_attendance')
    
    #view attendance
    attendance_records = StudentAttendance.objects.filter(faculty=faculty).order_by('-date') 
    for attendance in attendance_records:
        attendance.filename = os.path.basename(attendance.file.name)  
         
    return render(request, 'Faculty/faculty_attendance.html', {'attendance_by_year': attendance_by_year,'faculty':faculty,'form': form, 'attendance_records': attendance_records})
     
@faculty_required
def leave_request(request):
    faculty = request.user.faculty
    leave_requests = LeaveRequest.objects.filter(user=request.user)
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.user = request.user
            leave_request.status = 'Pending'  # Set status to 'Pending'
            leave_request.save()
            return redirect('leave_request')  
    else:
        form = LeaveRequestForm()
    
    user = request.user
    faculty_leaves = FacultyLeave.objects.all()
    months_order = list(calendar.month_name)[1:] 
    valid_faculty_leaves = [leave for leave in faculty_leaves if leave.month in months_order]
    faculty_leaves = sorted(valid_faculty_leaves, key=lambda leave: months_order.index(leave.month))
    leave_data = []
    for faculty_leave in faculty_leaves:
        try:
            month_num = datetime.strptime(faculty_leave.month, '%B').month
        except ValueError as e:
            continue
        leaves_requests = LeaveRequest.objects.filter(
            user=user,
            start_date__month=month_num,
            status='Approved'
        )
        leaves_taken = sum((request.end_date - request.start_date).days + 1 for request in leaves_requests)
        remaining_leaves = faculty_leave.faculty_leaves - leaves_taken

        leave_data.append({
            'month': faculty_leave.month,
            'allocated_leaves': faculty_leave.faculty_leaves,
            'leaves_taken': leaves_taken,
            'remaining_leaves': remaining_leaves
        })
    return render(request, 'Faculty/leave_request.html',
                   {"faculty": faculty,
                    'form': form, 
                    'leave_requests': leave_requests,
                    'leave_data':leave_data
                    })

@faculty_required
def faculty_timetable(request):
    faculty = request.user.faculty
    timetable = FacultyTimetable.objects.filter(faculty=faculty).last()
    timetable_data = []
    if timetable:
        file_path = timetable.file.path
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    timetable_data.append(row)
        elif file_path.endswith('.xlsx'):
            timetable_df = pd.read_excel(file_path)
            timetable_data = timetable_df.to_dict(orient='records')
        else:
            print("Unsupported file format for timetable:", file_path)
            timetable_data = None
    else:
        print("No Timetable instance found for faculty:", faculty.faculty_id)
        timetable_data = None
    return render(request, 'Faculty/timetable.html', {'faculty': faculty,'timetable_data':timetable_data})

@faculty_required
def faculty_payslip(request):
    faculty = request.user.faculty
    payslips = FacultySalary.objects.filter(faculty=faculty)
    return render(request, 'Faculty/salary.html', {'faculty':faculty,'payslips': payslips})

@faculty_required
def marks(request):
    # upload marks
    faculty = get_object_or_404(Faculty, user=request.user)
    form = MarksForm()
    classes = Class.objects.all()
    sections = Section.objects.all()
    subjects = Subject.objects.all()
    exam_types = [exam_type[0] for exam_type in Marks.EXAM_TYPES]
    error_message = None

    if request.method == 'POST':
        form = MarksForm(request.POST, request.FILES)
        if form.is_valid():
            class_id = request.POST.get('class_linked')
            section_id = request.POST.get('section_linked')
            subject_id = request.POST.get('subject')
            exam_type = request.POST.get('exam_type')
            unit_test_number = request.POST.get('unit_test_number')

            # Check if marks for the same class-section, subject, and exam type already exist
            existing_marks = Marks.objects.filter(
                class_linked_id=class_id,
                section_linked_id=section_id,
                subject_id=subject_id,
                exam_type=exam_type,
                faculty=faculty
            )
            if exam_type == 'Unit Test':
                existing_marks = existing_marks.filter(unit_test_number=unit_test_number)

            if existing_marks.exists():
                error_message = "Marks for this class-section, subject, and exam type have already been uploaded."
            else:
                student_marks = form.save(commit=False)
                student_marks.class_linked_id = class_id
                student_marks.section_linked_id = section_id
                student_marks.subject_id = subject_id
                student_marks.exam_type = exam_type
                if exam_type == 'Unit Test':
                    student_marks.unit_test_number = unit_test_number
                else:
                    student_marks.unit_test_number = None
                student_marks.faculty = faculty
                student_marks.save()
                return redirect('marks')
        else:
            logger.debug("Form is not valid")
            logger.debug("Form errors: %s", form.errors)

    
    sections_json = serialize('json', sections, fields=('id', 'name', 'class_linked'))
    subjects_json = serialize('json', subjects, fields=('id', 'subjectname', 'section_linked'))

    # view marks
    try:
        faculty = request.user.faculty
        student_marks = Marks.objects.filter(faculty=faculty)
    except AttributeError:
        student_marks = []
    exam_type_dict = {}
    for exam_type in exam_types:
        if exam_type == 'Unit Test':
            unit_test_dict = {}
            unit_test_numbers = student_marks.filter(exam_type='Unit Test').values_list('unit_test_number', flat=True).distinct()
            for unit_test_number in unit_test_numbers:
                unit_test_dict[unit_test_number] = student_marks.filter(exam_type='Unit Test', unit_test_number=unit_test_number)
            exam_type_dict['Unit Test'] = unit_test_dict
        else:
            exam_type_dict[exam_type] = student_marks.filter(exam_type=exam_type)

    return render(request, 'Faculty/marks.html', {
        'form':form,
        'classes': classes,
        'sections': sections,
        'subjects': subjects,
        'faculty': faculty,
        'exam_types':exam_types,
        'sections_json':sections_json ,
        'subjects_json':subjects_json,
        'exam_type_dict': exam_type_dict,
        'error_message': error_message
    })

@faculty_required
def faculty_bus_details(request):
    faculty = Faculty.objects.get(user=request.user) 
    faculty_location = faculty.location  
    matching_routes = BusRoute.objects.filter(pickup_locations__location=faculty_location)
    pickup_data = []
    dropoff_data = []
    # initialize values
    driver_name = None
    driver_number = None
    bus_number = None
    for route in matching_routes:
        driver_name = route.driver_name
        driver_number = route.driver_number
        bus_number = route.bus_number
        pickup_locations = PickupLocation.objects.filter(bus_route=route)
        dropoff_locations = DropoffLocation.objects.filter(bus_route=route)
        for pickup_location in pickup_locations:
            pickup_data.append({
                'pickup_location': pickup_location.location,
                'pickup_time': pickup_location.time,
                'driver_name': pickup_location.bus_route.driver_name,
                'bus_number': pickup_location.bus_route.bus_number,
                'driver_number': pickup_location.bus_route.driver_number
            })
        for dropoff_location in dropoff_locations:
            dropoff_data.append({
                'dropoff_location': dropoff_location.location,
                'dropoff_time': dropoff_location.time
            })

    context = {
        'faculty': faculty,
        'pickup_data': pickup_data,
        'dropoff_data': dropoff_data,
        'driver_name': driver_name,
        'driver_number': driver_number,
        'bus_number': bus_number
    }
    return render(request, 'Faculty/faculty_bus.html', context)

#Student
@student_required
def student_portal(request):
    student = request.user.students
    student_count = Students.objects.filter(
            class_name=student.class_name,
            section_linked=student.section_linked
        ).count()
    try:
        student = request.user.students
        subjects = Subject.objects.filter(class_linked=student.class_name, section_linked=student.section_linked)
        context = {"student": student, 'subjects':subjects, 'student_count': student_count,}
        return render(request, 'Student/student_portal.html', context)
    except Students.DoesNotExist:
        messages.error(request, "Your registration has been rejected.")
        return redirect('home')

@student_required
def student_classes(request):
    student = request.user.students
    context = {"student": student}
    return render(request, 'Student/classes.html', context)

@student_required
def student_attendance(request):
    student = request.user.students
    attendance_records = []
    files = StudentAttendance.objects.filter(
            class_assigned=student.class_name,
            section_assigned=student.section_linked
        ).order_by('date')
    for attendance_file in files:
        if attendance_file.file.name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(attendance_file.file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):  
                    if row[0] == student.name:
                        record = {
                            'date': attendance_file.date,
                            'status': row[1]
                        }
                        attendance_records.append(record)
            except InvalidFileException:
                continue  
            except Exception as e:
                continue
        elif attendance_file.file.name.endswith('.csv'):
            try:
                file_data = attendance_file.file.read().decode('utf-8').splitlines()
                reader = csv.reader(file_data)
                next(reader)  # Skip the header row
                for row in reader:
                    if row[0] == student.name:
                        record = {
                            'date': attendance_file.date,
                            'status': row[1]
                        }
                        attendance_records.append(record)
            except Exception as e:
                continue
            
    from collections import defaultdict
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'total_days': 0, 'days_absent': 0}))
    for record in attendance_records:
        year = record['date'].year
        month = record['date'].month
        monthly_data[year][month]['total_days'] += 1
        if record['status'] == 'Absent':
            monthly_data[year][month]['days_absent'] += 1
    attendance_by_year = {}
    for year, months in monthly_data.items():
        attendance_by_year[year] = []
        for month, data in months.items():
            total_days = data['total_days']
            days_absent = data['days_absent']
            days_present = total_days - days_absent
            attendance_percentage = (days_present / total_days) * 100 if total_days > 0 else 0
            attendance_by_year[year].append({
                'month': calendar.month_name[month],
                'total_days': total_days,
                'days_absent': days_absent,
                'days_present': days_present,
                'attendance_percentage': attendance_percentage
            })

    context = {"student": student,'attendance_by_year':attendance_by_year}
    return render(request, 'Student/attendance.html',context)

@student_required
def student_timetable(request):
    student = request.user.students
    timetable = StudentTimetable.objects.filter(
            class_linked=student.class_name,
            section_linked=student.section_linked).last()
    timetable_data = []
    if timetable:
        file_path = timetable.file.path
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    timetable_data.append(row)
        elif file_path.endswith('.xlsx'):
            timetable_df = pd.read_excel(file_path)
            timetable_data = timetable_df.to_dict(orient='records')
        else:
            print("Unsupported file format for timetable:", file_path)
            timetable_data = None
    else:
        print("No Timetable instance found for faculty:", student.student_id)
        timetable_data = None

    context = {
        'student':student,
        'timetable_data': timetable_data,
    }
    return render(request, 'Student/timetable.html', context)

@student_required
def exam_timetable(request):
    student = request.user.students
    exams = Exam.objects.filter(class_linked=student.class_name, section_linked=student.section_linked)
    
    timetable_data = []
    grouped_exams = {}

    for exam in exams:
        file_path = exam.file.path
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    date_str = row.get('Date', '')
                    if date_str:
                        date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                        row['Day'] = date_obj.strftime('%A')
                    timetable_data.append((row, exam.exam_type, exam.unit_test_number))
        elif file_path.endswith('.xlsx'):
            timetable_df = pd.read_excel(file_path)
            for _, row in timetable_df.iterrows():
                row_dict = row.to_dict()
                date_value = row_dict.get('Date', '')
                if isinstance(date_value, pd.Timestamp):
                    date_obj = date_value.to_pydatetime()
                elif isinstance(date_value, str):
                    date_obj = datetime.strptime(date_value, '%d/%m/%Y')
                else:
                    continue
                row_dict['Day'] = date_obj.strftime('%A')
                row_dict['Date'] = date_obj.strftime('%d/%m/%Y')
                timetable_data.append((row_dict, exam.exam_type, exam.unit_test_number))
        else:
            print("Unsupported file format for timetable:", file_path)
            continue

    if timetable_data:
        for row, exam_type, unit_test_number in timetable_data:
            if exam_type not in grouped_exams:
                grouped_exams[exam_type] = {}

            if exam_type == 'Unit Test':
                if unit_test_number not in grouped_exams[exam_type]:
                    grouped_exams[exam_type][unit_test_number] = []
                grouped_exams[exam_type][unit_test_number].append(row)
            else:
                if 'others' not in grouped_exams[exam_type]:
                    grouped_exams[exam_type]['others'] = []
                grouped_exams[exam_type]['others'].append(row)
        
    context = {
        "student": student,
        "grouped_exams": grouped_exams,
    }
    return render(request, 'Student/exam_timetable.html', context)

@student_required
def student_marks(request):
    student = request.user.students
    student_class = student.class_name
    student_section = student.section_linked
    marks_records = Marks.objects.filter(class_linked=student_class, section_linked=student_section)
    marks_data = {}
    for record in marks_records:
        file_path = record.file.path
        if file_path.endswith('.csv'):
            try:
                df = pd.read_csv(file_path)
                for index, row in df.iterrows():
                    if row['name'] == student.name:  # Assuming 'Name' column contains student names
                        exam_type = record.exam_type
                        subject = record.subject
                        marks = row['marks']

                        if exam_type not in marks_data:
                            marks_data[exam_type] = {}

                        if exam_type == 'Unit Test':
                            unit_test_number = record.unit_test_number
                            if unit_test_number not in marks_data[exam_type]:
                                marks_data[exam_type][unit_test_number] = []
                            marks_data[exam_type][unit_test_number].append({
                                'subject': subject,
                                'marks': marks
                            })
                        else:
                            if 'Other' not in marks_data[exam_type]:
                                marks_data[exam_type]['Other'] = []
                            marks_data[exam_type]['Other'].append({
                                'subject': subject,
                                'marks': marks
                            })
            except pd.errors.EmptyDataError:
                continue
            except Exception as e:
                continue

        elif file_path.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == student.name:  # Assuming student name is in the first column
                        exam_type = record.exam_type
                        subject = record.subject  # Assuming Subject is in the file
                        marks = row[1]    # Assuming Marks is in the file

                        if exam_type not in marks_data:
                            marks_data[exam_type] = {}
                        
                        if exam_type == 'Unit Test':
                            unit_test_number = record.unit_test_number
                            if unit_test_number not in marks_data[exam_type]:
                                marks_data[exam_type][unit_test_number] = []
                            marks_data[exam_type][unit_test_number].append({
                                'subject': subject,
                                'marks': marks
                            })
                        else:
                            if 'Other' not in marks_data[exam_type]:
                                marks_data[exam_type]['Other'] = []
                            marks_data[exam_type]['Other'].append({
                                'subject': subject,
                                'marks': marks
                            })
            except openpyxl.utils.exceptions.InvalidFileException:
                continue
            except Exception as e:
                continue
    return render(request, 'Student/student_marks.html',{
        'student':student,
        'marks_data':marks_data,})


#Parent
@parent_required
def parent_portal(request):
    try:
        parent = request.user.parents
        children = parent.children.all()
        subjects = []
        for child in children:
            child_subjects = Subject.objects.filter(class_linked=child.class_name, section_linked=child.section_linked)
            subjects.extend(child_subjects)
        return render(request, 'Parent/parent_portal.html',{'parent':parent, 'children':children, 'subjects':subjects})
    except Parents.DoesNotExist:
        messages.error(request, "Your registration has been rejected.")
        return redirect('home')

@parent_required    
def parent_classes(request):
    parent = request.user.parents
    children = parent.children.all()
    return render(request, 'Parent/classes.html',{'parent':parent, 'children':children})

@parent_required
def parent_attendance(request):
    parent = Parents.objects.get(user=request.user) 
    children = parent.children.all()
    attendance_records = [] 
    for child in children:
        attendance_files = StudentAttendance.objects.filter(
            class_assigned=child.class_name,
            section_assigned=child.section_linked
        ).order_by('date')
        for attendance_file in attendance_files:
            if attendance_file.file.name.endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(attendance_file.file)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] == child.name:
                            record = {
                                'date': attendance_file.date,
                                'status': row[1]
                            }
                            attendance_records.append(record)
                except InvalidFileException:
                    continue
                except Exception as e:
                    continue
            elif attendance_file.file.name.endswith('.csv'):
                try:
                    file_data = attendance_file.file.read().decode('utf-8').splitlines()
                    reader = csv.reader(file_data)
                    next(reader)  # Skip the header row
                    for row in reader:
                        if row[0] == child.name:
                            record = {
                                'date': attendance_file.date,
                                'status': row[1]
                            }
                            attendance_records.append(record)
                except Exception as e:
                    continue
            
    from collections import defaultdict
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'total_days': 0, 'days_absent': 0}))
    for record in attendance_records:
        year = record['date'].year
        month = record['date'].month
        monthly_data[year][month]['total_days'] += 1
        if record['status'] == 'Absent':
            monthly_data[year][month]['days_absent'] += 1
    attendance_by_year = {}
    for year, months in monthly_data.items():
        attendance_by_year[year] = []
        for month, data in months.items():
            total_days = data['total_days']
            days_absent = data['days_absent']
            days_present = total_days - days_absent
            attendance_percentage = (days_present / total_days) * 100 if total_days > 0 else 0
            attendance_by_year[year].append({
                'month': calendar.month_name[month],
                'total_days': total_days,
                'days_absent': days_absent,
                'days_present': days_present,
                'attendance_percentage': attendance_percentage
            })

    context = {'parent':parent, 'children':children,'attendance_by_year':attendance_by_year}
    return render(request, 'Parent/attendance.html',context)

@parent_required
def parent_timetable(request):
    parent = Parents.objects.get(user=request.user) 
    children = parent.children.all() 
    timetable_data = []
    for child in children:
        timetable = StudentTimetable.objects.filter(
            class_linked=child.class_name,
            section_linked=child.section_linked).last()
        if timetable:
            file_path = timetable.file.path
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        timetable_data.append(row)
            elif file_path.endswith('.xlsx'):
                timetable_df = pd.read_excel(file_path)
                timetable_data = timetable_df.to_dict(orient='records')
            else:
                print("Unsupported file format for timetable:", file_path)
                timetable_data = None
        else:
            print("No Timetable instance found for child:", child.student_id)
            timetable_data = None

    context = {
        'timetable_data': timetable_data,
        'children':children,
        'parent': parent,
    }
    return render(request, 'Parent/parent_timetable.html', context)

@parent_required
def parent_exam_timetable(request):
    parent = Parents.objects.get(user=request.user)
    children = parent.children.all()

    grouped_exams = {}

    for child in children:
        exams = Exam.objects.filter(class_linked=child.class_name, section_linked=child.section_linked)
        timetable_data = []

        for exam in exams:
            file_path = exam.file.path
            if file_path.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    for row in reader:
                        date_str = row.get('Date', '')
                        if date_str:
                            date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                            row['Day'] = date_obj.strftime('%A')
                        timetable_data.append((row, exam.exam_type, exam.unit_test_number))
            elif file_path.endswith('.xlsx'):
                timetable_df = pd.read_excel(file_path)
                for _, row in timetable_df.iterrows():
                    row_dict = row.to_dict()
                    date_value = row_dict.get('Date', '')
                    if isinstance(date_value, pd.Timestamp):
                        date_obj = date_value.to_pydatetime()
                    elif isinstance(date_value, str):
                        date_obj = datetime.strptime(date_value, '%d/%m/%Y')
                    else:
                        continue
                    row_dict['Day'] = date_obj.strftime('%A')
                    row_dict['Date'] = date_obj.strftime('%d/%m/%Y')
                    timetable_data.append((row_dict, exam.exam_type, exam.unit_test_number))
            else:
                print("Unsupported file format for timetable:", file_path)
                continue

        if timetable_data:
            for row, exam_type, unit_test_number in timetable_data:
                if exam_type not in grouped_exams:
                    grouped_exams[exam_type] = {}

                if exam_type == 'Unit Test':
                    if unit_test_number not in grouped_exams[exam_type]:
                        grouped_exams[exam_type][unit_test_number] = []
                    grouped_exams[exam_type][unit_test_number].append(row)
                else:
                    if 'others' not in grouped_exams[exam_type]:
                        grouped_exams[exam_type]['others'] = []
                    grouped_exams[exam_type]['others'].append(row)

    context = {
        'children': children,
        'parent': parent,
        "grouped_exams": grouped_exams,
    }
    return render(request, 'Parent/parent_exam_timetable.html', context)

@parent_required
def parent_marks(request):
    parent = Parents.objects.get(user=request.user) 
    children = parent.children.all()
    marks_data = {}
    for child in children:
        student_class = child.class_name
        student_section = child.section_linked
        marks_records = Marks.objects.filter(class_linked=student_class, section_linked=student_section)
        for record in marks_records:
            file_path = record.file.path
            if file_path.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path)
                    for index, row in df.iterrows():
                        if row['name'] == child.name:  # Assuming 'name' column contains student names
                            exam_type = record.exam_type
                            subject = record.subject
                            marks = row['marks']

                            if exam_type not in marks_data:
                                marks_data[exam_type] = {}

                            if exam_type == 'Unit Test':
                                unit_test_number = record.unit_test_number
                                if unit_test_number not in marks_data[exam_type]:
                                    marks_data[exam_type][unit_test_number] = []
                                marks_data[exam_type][unit_test_number].append({
                                    'subject': subject,
                                    'marks': marks
                                })
                            else:
                                if 'Other' not in marks_data[exam_type]:
                                    marks_data[exam_type]['Other'] = []
                                marks_data[exam_type]['Other'].append({
                                    'subject': subject,
                                    'marks': marks
                                })
                except pd.errors.EmptyDataError:
                    continue
                except Exception as e:
                    continue

            elif file_path.endswith('.xlsx'):
                try:
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] == child.name:  # Assuming student name is in the first column
                            exam_type = record.exam_type
                            subject = record.subject
                            marks = row[1]

                            if exam_type not in marks_data:
                                marks_data[exam_type] = {}

                            if exam_type == 'Unit Test':
                                unit_test_number = record.unit_test_number
                                if unit_test_number not in marks_data[exam_type]:
                                    marks_data[exam_type][unit_test_number] = []
                                marks_data[exam_type][unit_test_number].append({
                                    'subject': subject,
                                    'marks': marks
                                })
                            else:
                                if 'Other' not in marks_data[exam_type]:
                                    marks_data[exam_type]['Other'] = []
                                marks_data[exam_type]['Other'].append({
                                    'subject': subject,
                                    'marks': marks
                                })
                except openpyxl.utils.exceptions.InvalidFileException:
                    continue
                except Exception as e:
                    continue

    context = {
        'children': children,
        'parent': parent,
        'marks_data': marks_data
    }
    return render(request, 'Parent/parent_marks.html',context)

@parent_required
def parent_bus_details(request):
    parent = Parents.objects.get(user=request.user) 
    children = parent.children.all()
    bus_number = "N/A"
    driver_name = "N/A"
    driver_number = "N/A"
    parent_location = parent.location  
    matching_routes = BusRoute.objects.filter(pickup_locations__location=parent_location)
    pickup_data = []
    dropoff_data = []
    for route in matching_routes:
        driver_name = route.driver_name
        driver_number = route.driver_number
        bus_number = route.bus_number
        pickup_locations = PickupLocation.objects.filter(bus_route=route)
        dropoff_locations = DropoffLocation.objects.filter(bus_route=route)
        for pickup_location in pickup_locations:
            pickup_data.append({
                'pickup_location': pickup_location.location,
                'pickup_time': pickup_location.time,
                'driver_name': pickup_location.bus_route.driver_name,
                'bus_number': pickup_location.bus_route.bus_number,
                'driver_number': pickup_location.bus_route.driver_number
            })
        for dropoff_location in dropoff_locations:
            dropoff_data.append({
                'dropoff_location': dropoff_location.location,
                'dropoff_time': dropoff_location.time
            })

    context = {
        'parent': parent,
        'children': children,
        'pickup_data': pickup_data,
        'dropoff_data': dropoff_data,
        'driver_name': driver_name,
        'driver_number': driver_number,
        'bus_number': bus_number
    }
    return render(request, 'Parent/parent_bus.html', context)

#Payment
client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

@parent_required
def parent_fees(request):
    parent = Parents.objects.get(user=request.user)
    children = parent.children.all()
    fees_summary = []
    for child in children:
        child_fees_summary = {
            'student': child.name,
            'student_id': child.student_id,
            'fees_details': [],
            'payments': Payment.objects.filter(student=child)
        }

        # Tuition Fees
        tuition_fees = TuitionFees.objects.filter(class_linked=child.class_name).first()
        tuition_payments = Payment.objects.filter(student=child, fees_type='Tuition Fees')
        total_tuition_paid = sum(payment.amount for payment in tuition_payments)
        fees_to_be_paid = tuition_fees.amount - total_tuition_paid if tuition_fees else 0

        child_fees_summary['fees_details'].append({
            'fees_type': 'Tuition Fees',
            'total_fees': tuition_fees.amount if tuition_fees else 'N/A',
            'fees_paid': total_tuition_paid,
            'fees_to_be_paid': fees_to_be_paid
        })

        # Bus Fees
        bus_fees = BusFees.objects.filter(location=child.location).first()
        bus_payments = Payment.objects.filter(student=child, fees_type='Bus Fees')
        total_bus_paid = sum(payment.amount for payment in bus_payments)
        fees_to_be_paid = bus_fees.amount - total_bus_paid if bus_fees else 0

        child_fees_summary['fees_details'].append({
            'fees_type': 'Bus Fees',
            'total_fees': bus_fees.amount if bus_fees else 'N/A',
            'fees_paid': total_bus_paid,
            'fees_to_be_paid': fees_to_be_paid
        })

        # Mess Fees
        mess_fees = MessFees.objects.first()
        mess_payments = Payment.objects.filter(student=child, fees_type='Mess Fees')
        total_mess_paid = sum(payment.amount for payment in mess_payments)
        fees_to_be_paid = mess_fees.amount - total_mess_paid if mess_fees else 0

        child_fees_summary['fees_details'].append({
            'fees_type': 'Mess Fees',
            'total_fees': mess_fees.amount if mess_fees else 'N/A',
            'fees_paid': total_mess_paid,
            'fees_to_be_paid': fees_to_be_paid
        })

        fees_summary.append(child_fees_summary)
    return render(request, 'Parent/fees.html', {'parent': parent, 'children': children, 'fees_summary': fees_summary})

def create_order(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        fees_type = request.POST.get('fees_type')
        amount = request.POST.get('amount')
        student = Students.objects.get(student_id=student_id)

        # Ensure amount is correctly formatted and convert to paise
        try:
            amount = float(amount)
            amount = int(amount * 100)
        except ValueError:
            return JsonResponse({'status': 'invalid amount'})

        order = client.order.create({
            'amount': amount,
            'currency': 'INR',
            'payment_capture': '1'
        })

        context = {
            'order_id': order['id'],
            'amount': amount / 100,
            'key_id': settings.RAZORPAY_KEY_ID,
            'student_id': student_id,
            'fees_type': fees_type,
            'student_name': student.name,
            'class_name': student.class_name.name,
            'section_name': student.section_linked.name,
        }
        return render(request, 'Parent/payment.html', context)
    return redirect('parent_fees')

@csrf_exempt
def payment_callback(request):
    if request.method == 'POST':
        received_data = request.POST
        order_id = received_data.get('razorpay_order_id')
        payment_id = received_data.get('razorpay_payment_id')
        payment_signature = received_data.get('razorpay_signature')
        student_id = received_data.get('student_id')
        fees_type = received_data.get('fees_type')
        amount = received_data.get('amount')

        try:
            amount = float(amount)  # Ensure amount is converted to float
            student = Students.objects.get(student_id=student_id)
            client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
            payment = client.payment.fetch(payment_id)
            payment_status = payment.get('status', 'unknown') 
            payment = Payment.objects.create(
                student=student,
                fees_type=fees_type,
                amount=amount,
                receipt_number=payment_id,
                payment_type='Razorpay',
                class_linked=student.class_name,  
                section_linked=student.section_linked  
            )

            # Generate invoice
            context = {
                'student': student,
                'payment': payment,
                'payment_status': get_payment_status_display(payment_status),
            }
            invoice_pdf = render_to_pdf('Parent/invoice.html', context)
            if invoice_pdf:
                payment.invoice.save(f'invoice_{payment_id}.pdf', invoice_pdf)

            messages.success(request, 'Payment successful and recorded.')
        except Students.DoesNotExist:
            messages.error(request, 'Student not found. Payment not recorded.')
        except Exception as e:
            messages.error(request, f'Error recording payment: {e}')

        return redirect('parent_fees')

def get_payment_status_display(status):
    status_mapping = {
        'captured': 'Success',
        'failed': 'Failed',
        'pending': 'Pending',
        'authorized': 'Authorized',
        'cancelled': 'Cancelled'
    }
    return status_mapping.get(status, 'Unknown')        